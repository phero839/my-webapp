# -*- coding: utf-8 -*-
"""핵심 변환 로직"""
import io, json, os, re
from typing import Dict, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pdfplumber

NTS = {
    "cash":"1. 현금과 예금", "related_ar":"2. 특수관계인에 대한 매출채권",
    "other_ar":"3. 기타매출채권", "inventory":"4. 재고자산",
    "securities":"5. 유가증권", "investments":"6. 투자유가증권",
    "related_loans_recv":"7. 특수관계인에 대한 대여금", "other_loans_recv":"8. 기타대여금",
    "land_buildings":"1) 토지 및 건축물", "machinery_vehicles":"2) 기계장치, 차량운반구",
    "other_fixed":"3) 기타유형자산", "intangibles":"10. 무형자산",
    "other_assets":"11. 위 분류과목 이외 자산",
    "related_ap":"1. 특수관계자에 대한 매입채무", "other_ap":"2. 기타매입채무",
    "related_borrowings":"3. 특수관계자에 대한 차입금", "other_borrowings":"4. 기타차입금",
    "accrued":"5. 미지급금", "other_liabilities":"6. 위분류과목이외부채",
    "capital":"1. 자본금", "capital_surplus":"1) 자본잉여금",
    "retained_earnings":"2) 이익잉여금", "other_equity":"3) 기타",
    "related_sales":"1. 특수관계자에 대한 매출", "other_sales":"2. 기타 매출",
    "related_purchases":"1. 특수관계자로부터 매입", "other_purchases":"2. 기타 매입",
    "salary_parent":"1. 급여(모회사파견직원)", "salary_other":"2. 급여(기타)",
    "rent":"3. 임차료", "rnd":"4. 연구개발비",
    "bad_debt":"5. 대손상각비", "other_sga":"6. 기타판매비와관리비",
    "interest_income":"1. 이자수익", "dividend_income":"2. 배당금수익",
    "debt_forgiveness":"3. 채무면제익", "other_non_op_income":"4. 기타영업외수익",
    "interest_expense":"1. 이자비용", "other_non_op_expense":"2. 기타영업외비용",
    "income_tax":"법인세비용",
}

# ── 파싱 헬퍼 ─────────────────────────────────────────────────

def _parse_trial_balance(ws, target):
    """Japanese trial balance: col[0]=account, col[4]=current balance (当期残高)."""
    for row in ws.iter_rows(values_only=True):
        if not row or not isinstance(row[0], str):
            continue
        name = row[0].strip()
        if not name or name.startswith('['):
            continue
        if '合計' in name or name.endswith('計'):
            continue
        val = row[4] if len(row) > 4 else None
        if not isinstance(val, (int, float)) or val == 0:
            continue
        target[name] = val


def _parse_korean_bs(ws, bs):
    """Korean two-column BS: [asset_name, asset_val, liability_name, liability_val, ...]."""
    SKIP_KW = ('총계', '합계', 'total')
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Left pair (assets)
        n0 = row[0].strip() if isinstance(row[0], str) else None
        v1 = row[1] if len(row) > 1 and isinstance(row[1], (int, float)) else None
        if n0 and v1 is not None and not any(k in n0.lower() for k in SKIP_KW):
            bs[n0] = v1
        # Right pair (liabilities / equity)
        n2 = row[2].strip() if len(row) > 2 and isinstance(row[2], str) else None
        v3 = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) else None
        if n2 and v3 is not None and not any(k in n2.lower() for k in SKIP_KW):
            bs[n2] = v3


def _parse_korean_pl(ws, pl):
    """Korean PL: name in col[3], value in col[4]. Keep section totals + income tax."""
    for row in ws.iter_rows(values_only=True):
        name_cell = row[3] if len(row) > 3 else None
        val_cell  = row[4] if len(row) > 4 else None
        if not isinstance(name_cell, str):
            continue
        name = name_cell.strip()
        if not name or name.startswith('<'):
            break  # date-section marker → stop
        if not isinstance(val_cell, (int, float)):
            continue
        lo = name.lower()
        if lo.startswith('total') or 'income tax' in lo or '법인세' in lo:
            pl[name] = val_cell


def _parse_generic(ws, target):
    """Generic: first text cell = key, last numeric cell = value."""
    for row in ws.iter_rows(values_only=True):
        texts = [str(c).strip() for c in row if isinstance(c, str) and str(c).strip()]
        nums  = [c for c in row if isinstance(c, (int, float))]
        if texts and nums and not texts[0].startswith("="):
            target[texts[0]] = nums[-1]

# ── 파싱 ──────────────────────────────────────────────────────

def parse_excel(stream, bs_hint=None, pl_hint=None) -> Tuple[Dict, Dict]:
    wb = openpyxl.load_workbook(stream, data_only=True)
    bs, pl = {}, {}
    for name in wb.sheetnames:
        ws  = wb[name]
        lo  = name.lower()
        target = None

        if bs_hint and name == bs_hint:
            target = bs
        elif pl_hint and name == pl_hint:
            target = pl
        elif not (bs_hint or pl_hint):
            if any(k in lo for k in ["bs","balance","asset","재무상태","대차"]) or name == '貸':
                target = bs
            elif any(k in lo for k in ["pl","p&l","income","profit","loss","손익"]) or name == '損':
                target = pl

        if target is None:
            if len(wb.sheetnames) == 1:
                target = bs  # 단일 시트 파일 → BS로 처리 (홍콩법인 등)
            else:
                continue

        # 시트 형식에 맞는 파서 선택
        if name in ('貸', '損'):
            _parse_trial_balance(ws, target)
        elif '재무상태표' in name:
            _parse_korean_bs(ws, bs)
        elif '손익계산서' in name:
            _parse_korean_pl(ws, pl)
        elif len(wb.sheetnames) == 1:
            _parse_korean_bs(ws, bs)  # 단일 시트 = 한국형 2컬럼 BS
        else:
            _parse_generic(ws, target)
    return bs, pl


_BS_CODES = frozenset({
    "cash","related_ar","other_ar","inventory","securities","investments",
    "related_loans_recv","other_loans_recv","land_buildings","machinery_vehicles",
    "other_fixed","intangibles","other_assets","related_ap","other_ap",
    "related_borrowings","other_borrowings","accrued","other_liabilities",
    "capital","capital_surplus","retained_earnings","other_equity",
})

def _split_pdf_accounts(data: Dict) -> Tuple[Dict, Dict]:
    """단일 PDF에서 파싱된 계정을 BS/PL로 자동 분류."""
    bs, pl = {}, {}
    for name, val in data.items():
        r_bs = keyword_map(name, "bs")
        if r_bs == "skip":
            continue
        if r_bs in _BS_CODES:
            bs[name] = val
        else:
            r_pl = keyword_map(name, "pl")
            if r_pl != "skip":
                pl[name] = val
    return bs, pl


def parse_pdf(stream) -> Dict:
    data = {}
    with pdfplumber.open(stream) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = line.strip()
                m = re.match(r"^(.+?)\s+(\([\d,]+\.?\d*\)|[-]?\d[\d,]*\.?\d*)$", line)
                if m:
                    name = m.group(1).strip()
                    raw  = m.group(2).replace(",", "")
                    if raw.startswith("(") and raw.endswith(")"):
                        raw = "-" + raw[1:-1]
                    try: data[name] = float(raw)
                    except: pass
    return data


def detect_currency_year(stream, fmt: str) -> Tuple[str, int]:
    cur, year = "", 2025
    currencies = [
        "USD","JPY","EUR","GBP","CAD","CHF","AUD","NZD","CNH","CNY","HKD",
        "TWD","MNT","KZT","THB","SGD","IDR","MYR","PHP","VND","BND",
        "INR","PKR","BDT","KHR","MOP","NPR","LKR","UZS","MMK",
        "MXN","BRL","ARS","CLP","COP",
        "SEK","DKK","NOK","RUB","HUF","PLN","CZK","RON",
        "SAR","QAR","ILS","JOD","KWD","BHD","AED","TRY","OMR",
        "ZAR","EGP","KES","LYD","ETB","FJD",
    ]
    if fmt == "excel":
        wb = openpyxl.load_workbook(stream, data_only=True)
        for sname in wb.sheetnames:
            m = re.search(r"20(\d{2})", sname)
            if m: year = int("20"+m.group(1))
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        for c in currencies:
                            if c in cell.upper() and not cur: cur = c
    else:
        with pdfplumber.open(stream) as pdf:
            text = pdf.pages[0].extract_text() or ""
            for line in text.split("\n")[:15]:
                m = re.search(r"20(\d{2})", line)
                if m: year = int("20"+m.group(1))
                for c in currencies:
                    if c in line.upper() and not cur: cur = c
    return (cur or "USD"), year

# ── 매핑 ──────────────────────────────────────────────────────

def _norm(n):
    s = re.sub(r"([a-z])([A-Z])", r"\1 \2", n)
    s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", s)
    return re.sub(r"[_\-&/]", " ", s).lower()

def keyword_map(name: str, ctx: str) -> Optional[str]:
    n = _norm(name); raw = name.lower()
    def has(*kws): return any(k in n or k in raw for k in kws)

    # ── PL 소계 행 우선 매핑 (generic skip 전) ───────────────────
    if ctx == "pl":
        if has("total income", "총수익"):                        return "other_sales"
        if has("total cogs", "total cost of goods"):             return "other_purchases"
        if has("total expense", "total operating expense"):      return "other_sga"
        # 일본어 PL 합계 행 → skip
        if has("当期純損益", "経常損益", "売上総損益", "税引前", "営業損益"): return "skip"
        # 영문 net/gross 합계 → skip
        if has("net income", "net profit", "net loss", "net other income",
               "net ordinary", "gross profit"): return "skip"

    # ── BS equity 에서 당기순손익 → 이익잉여금 ──────────────────
    if ctx != "pl":
        if has("net income", "net profit", "net loss",
               "当期純損益", "当期純利益", "当期純損失"):          return "retained_earnings"

    # ── Generic skip ──────────────────────────────────────────────
    if has("total ","합계","소계","subtotal",
           "gross profit","net ordinary","ordinary income","trading income total",
           "cost of sales total","operating expenses total","balance sheet","profit and loss"):
        return "skip"
    if re.match(r"^[\d\s\-]+$", n): return "skip"

    # ── PL 컨텍스트 ───────────────────────────────────────────────
    if ctx == "pl":
        if has("sales","revenue","trading income","service income","marketing sales","pr service","매출",
               "shipping","delivery income","refund","seller discount"):
            return "related_sales" if has("related","intercompany","parent","subsidiary") else "other_sales"
        if has("cost of goods","costofgoods","cost of sales","costofsales","cogs","매출원가",
               "仕入","棚卸","売上原価"):                         return "other_purchases"
        if has("salary","salaries","wage","payroll","compensation","급여","給与","給料"):
            return "salary_parent" if has("dispatch","parent","모회사","派遣") else "salary_other"
        if has("rent","lease expense","operating lease","임차료","家賃","賃借"): return "rent"
        if has("research","development","r&d","연구개발","研究"):  return "rnd"
        if has("bad debt","doubtful","대손"):                      return "bad_debt"
        if has("interest income","이자수익","受取利息"):           return "interest_income"
        if has("interest expense","finance cost","이자비용","支払利息"): return "interest_expense"
        if has("dividend","배당","配当"):
            return "dividend_income" if has("income","receiv") else "other_non_op_income"
        if has("income tax","tax expense","tax provision","법인세","法人税"): return "income_tax"
        if has("exchange gain","fx gain","환차익"):                return "other_non_op_income"
        if has("exchange loss","fx loss","환차손"):                return "other_non_op_expense"
        return "other_sga"

    # ── BS 컨텍스트 ───────────────────────────────────────────────
    # 현금·예금 (한국어·일본어·영어)
    if has("cash","checking","saving","현금","예금","통장",
           "現金","普通預金","当座預金","預金"):                   return "cash"
    if has("bank","account") and not has("payable","liabilit","bank fee","bank charge","bank service"):
        return "cash"
    if re.search(r"\b\d{3,}\b", name) and ctx != "pl":         return "cash"

    # 매출채권
    if has("account receivable","accounts receivable","trade receivable","매출채권","売掛金","受取手形"):
        return "related_ar" if has("related","intercompany","parent","subsidiary") else "other_ar"
    if has("other receivable","receivable","미수금"):            return "other_ar"

    # 재고
    if has("inventory","재고","棚卸","商品","製品"):              return "inventory"

    # 유가증권·투자
    if has("marketable securities","short term investment"):     return "securities"
    if has("long term investment","equity investment","투자유가","投資有価証券"): return "investments"

    # 유형자산
    if has("land","building","real estate","property plant","土地","建物","建設仮勘定"): return "land_buildings"
    if has("machinery","equipment","vehicle","furniture","fixture","leasehold improvement",
           "right of use","rou asset","機械","車両","器具","備品"): return "other_fixed"
    if has("intangible","goodwill","patent","trademark","のれん","無形"): return "intangibles"
    if has("prepaid","deposit","deferred charge","advance payment","선급","보증금"): return "other_assets"

    # 매입채무
    if has("account payable","accounts payable","trade payable","매입채무","買掛金","支払手形"):
        return "related_ap" if has("related","intercompany","parent") else "other_ap"

    # 차입금
    if has("loan from member","member loan","shareholder loan","loan from shareholder",
           "loan from owner","loan from parent"): return "related_borrowings"
    if has("bank loan","note payable","bond payable","borrowing","차입금","借入"):
        return "other_borrowings"

    # 미지급금·기타부채
    if has("other payable","accrued","interest payable","wages payable","salary payable","미지급"):
        return "accrued"
    if has("gst","vat","tax payable","deferred revenue","lease liabilit",
           "未払法人税","未払消費税","未払金"):                    return "other_liabilities"

    # 자본
    if re.search(r"\b(member|owner)\s*[a-z0-9]\b", n):         return "capital"
    if has("member contribution","owner contribution","share capital","paid in capital",
           "common stock","opening balance equity","자본금","資本金"): return "capital"
    if has("owner","member") and has("capital","contribution","share"): return "capital"
    if has("additional paid","share premium","capital surplus","자본잉여","資本準備金"): return "capital_surplus"
    if has("retained earning","accumulated deficit","current year earning",
           "이익잉여금","繰越利益剰余金","利益剰余金"):            return "retained_earnings"

    return None


def map_with_claude(bs, pl, company, currency) -> Dict[str, str]:
    try:
        import anthropic
        key = os.environ.get("ANTHROPIC_API_KEY","")
        if not key: return {}
        client = anthropic.Anthropic(api_key=key)
    except ImportError:
        return {}

    prompt = f"""Korean tax accountant: map financial items from "{company}" ({currency}) to NTS codes.

BS: {json.dumps(bs, ensure_ascii=False)}
PL: {json.dumps(pl, ensure_ascii=False)}

Codes: cash,related_ar,other_ar,inventory,securities,investments,related_loans_recv,other_loans_recv,
land_buildings,machinery_vehicles,other_fixed,intangibles,other_assets,related_ap,other_ap,
related_borrowings,other_borrowings,accrued,other_liabilities,capital,capital_surplus,
retained_earnings,other_equity,related_sales,other_sales,related_purchases,other_purchases,
salary_parent,salary_other,rent,rnd,bad_debt,other_sga,interest_income,dividend_income,
debt_forgiveness,other_non_op_income,interest_expense,other_non_op_expense,income_tax,skip

Rules:
- totals/subtotals/net income/gross profit → skip
- "Total Income","총수익" → other_sales
- "Total COGS","Total Cost of Goods" → other_purchases
- "Total Expense","Total Operating Expense" → other_sga
- "Income Tax Provision/Payable" in PL → income_tax
- "Income Tax Payable" in BS → other_liabilities
- "Loan from Member"→related_borrowings
- "Member 1/2"/"Owner A"/"Opening Balance Equity"→capital
- bank account names with numbers→cash
- "Net Income"/"当期純損益金額" in BS equity section→retained_earnings
- "Retained Earnings"+"Net Income" both in BS equity→both map to retained_earnings (summed)
- "Account Payable"/"Other Payable*" for service company → accrued (미지급금)
- "Other Receivables_HQ" → other_assets
- Japanese: 普通預金→cash, 資本金→capital, 未払法人税等→other_liabilities, 法人税等(PL)→income_tax
- all unmatched PL expenses→other_sga

Return ONLY JSON: {{"name":"code",...}}"""
    try:
        r = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=3000,
                                   messages=[{"role":"user","content":prompt}])
        raw = re.sub(r"^```(?:json)?\n?|\n?```$","",r.content[0].text.strip())
        return json.loads(raw)
    except:
        return {}


def build_mapping(bs, pl, company, currency):
    m = map_with_claude(bs, pl, company, currency)
    for acct in bs:
        if acct not in m: m[acct] = keyword_map(acct,"bs") or "skip"
    for acct in pl:
        if acct not in m: m[acct] = keyword_map(acct,"pl") or "skip"
    return m


def apply_mapping(bs, pl, mapping) -> Dict:
    result = {k: None for k in NTS}
    all_data = {**bs, **pl}
    for acct, code in mapping.items():
        if code == "skip" or code not in result: continue
        v = all_data.get(acct)
        if v is not None:
            result[code] = (result[code] or 0) + v
    return result

# ── Excel 생성 ─────────────────────────────────────────────────

BOLD       = Font(bold=True)
CENTER     = Alignment(horizontal="center")
HDR_FILL   = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")

def _c(ws, row, col, val, bold=False, fill=None, align=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    if bold:  cell.font  = BOLD
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt
    return cell

def write_sheet(wb, company, sheet_name, d, eoy, avg, prior):
    ws = wb.create_sheet(title=sheet_name)
    v  = lambda k: (d[k] if d.get(k) else None)

    _c(ws,1,2,company,bold=True)
    _c(ws,2,3,"기말",align=CENTER,fill=HDR_FILL)
    _c(ws,2,6,"연평균",align=CENTER,fill=HDR_FILL)
    _c(ws,3,2,"환율",bold=True,fill=HDR_FILL)
    ws.cell(3,3).value=eoy; ws.cell(3,6).value=avg
    for col,lbl in [(2,"1. 현지법인 요약재무상태표"),(3,"외화"),(4,"원화"),
                    (5,"2. 현지법인 요약손익계산서"),(6,"외화"),(7,"원화"),
                    (9,"3. 현지법인 이잉/결손 계산서"),(10,"원화")]:
        _c(ws,4,col,lbl,bold=True,fill=HDR_FILL,align=CENTER)

    _c(ws,5,2,"I. 자산총계",bold=True,fill=TOTAL_FILL)
    ws.cell(5,3).value="=SUM(C6:C14)+SUM(C18:C19)"; ws.cell(5,4).value="=SUM(D6:D14)+SUM(D18:D19)"
    for row,key,lbl in [(6,"cash","1. 현금과 예금"),(7,"related_ar","2. 특수관계인에 대한 매출채권"),
        (8,"other_ar","3. 기타매출채권"),(9,"inventory","4. 재고자산"),
        (10,"securities","5. 유가증권"),(11,"investments","6. 투자유가증권"),
        (12,"related_loans_recv","7. 특수관계인에 대한 대여금"),(13,"other_loans_recv","8. 기타대여금")]:
        _c(ws,row,2,lbl); ws.cell(row,3).value=v(key); ws.cell(row,4).value=f"=ROUND(C{row}*$C$3,0)"
    _c(ws,14,2,"9. 유형자산",bold=True); ws.cell(14,3).value="=SUM(C15:C17)"; ws.cell(14,4).value="=SUM(D15:D17)"
    for row,key,lbl in [(15,"land_buildings","  1) 토지 및 건축물"),
        (16,"machinery_vehicles","  2) 기계장치, 차량운반구"),(17,"other_fixed","  3) 기타유형자산")]:
        _c(ws,row,2,lbl); ws.cell(row,3).value=v(key); ws.cell(row,4).value=f"=ROUND(C{row}*$C$3,0)"
    _c(ws,18,2,"10. 무형자산"); ws.cell(18,3).value=v("intangibles"); ws.cell(18,4).value="=ROUND(C18*$C$3,0)"
    _c(ws,19,2,"11. 위 분류과목 이외 자산"); ws.cell(19,3).value=v("other_assets"); ws.cell(19,4).value="=ROUND(C19*$C$3,0)"

    _c(ws,20,2,"II. 부채총계",bold=True,fill=TOTAL_FILL)
    ws.cell(20,3).value="=SUM(C21:C26)"; ws.cell(20,4).value="=SUM(D21:D26)"
    for row,key,lbl in [(21,"related_ap","1. 특수관계자에 대한 매입채무"),(22,"other_ap","2. 기타매입채무"),
        (23,"related_borrowings","3. 특수관계자에 대한 차입금"),(24,"other_borrowings","4. 기타차입금"),
        (25,"accrued","5. 미지급금"),(26,"other_liabilities","6. 위분류과목이외부채")]:
        _c(ws,row,2,lbl); ws.cell(row,3).value=v(key); ws.cell(row,4).value=f"=ROUND(C{row}*$C$3,0)"

    _c(ws,27,2,"III. 자본금총계",bold=True,fill=TOTAL_FILL)
    ws.cell(27,3).value="=SUM(C28:C29)"; ws.cell(27,4).value="=D5-D20"
    _c(ws,28,2,"1. 자본금"); ws.cell(28,3).value=v("capital"); ws.cell(28,4).value="=ROUND(C28*$C$3,0)"
    _c(ws,29,2,"2. 기타자본금"); ws.cell(29,3).value="=SUM(C30:C32)"; ws.cell(29,4).value="=SUM(D30:D32)"
    _c(ws,30,2,"  1) 자본잉여금"); ws.cell(30,3).value=v("capital_surplus"); ws.cell(30,4).value="=ROUND(C30*$C$3,0)"
    _c(ws,31,2,"  2) 이익잉여금"); ws.cell(31,3).value=v("retained_earnings"); ws.cell(31,4).value="=J10"
    _c(ws,32,2,"  3) 기타"); ws.cell(32,4).value="=D27-D28-D30-D31"
    ws.cell(33,3).value="=C27=C5-C20"; ws.cell(33,4).value="=D27=D5-D20"

    _c(ws,5,5,"I. 매출액",bold=True,fill=TOTAL_FILL)
    ws.cell(5,6).value="=SUM(F6:F7)"; ws.cell(5,7).value="=SUM(G6:G7)"
    for row,key,lbl in [(6,"related_sales","1. 특수관계자에 대한 매출"),(7,"other_sales","2. 기타 매출")]:
        _c(ws,row,5,lbl); ws.cell(row,6).value=v(key); ws.cell(row,7).value=f"=ROUND(F{row}*$F$3,0)"
    _c(ws,8,5,"II. 매출원가",bold=True,fill=TOTAL_FILL)
    ws.cell(8,6).value="=SUM(F9:F10)"; ws.cell(8,7).value="=SUM(G9:G10)"
    for row,key,lbl in [(9,"related_purchases","1. 특수관계자로부터 매입"),(10,"other_purchases","2. 기타 매입")]:
        _c(ws,row,5,lbl); ws.cell(row,6).value=v(key); ws.cell(row,7).value=f"=ROUND(F{row}*$F$3,0)"
    _c(ws,11,5,"III. 매출총손익",bold=True,fill=TOTAL_FILL); ws.cell(11,6).value="=F5-F8"; ws.cell(11,7).value="=G5-G8"
    _c(ws,12,5,"IV. 판매비와 일반관리비",bold=True,fill=TOTAL_FILL)
    ws.cell(12,6).value="=SUM(F13:F18)"; ws.cell(12,7).value="=SUM(G13:G18)"
    for row,key,lbl in [(13,"salary_parent","1. 급여(모회사파견직원)"),(14,"salary_other","2. 급여(기타)"),
        (15,"rent","3. 임차료"),(16,"rnd","4. 연구개발비"),
        (17,"bad_debt","5. 대손상각비"),(18,"other_sga","6. 기타판매비와관리비")]:
        _c(ws,row,5,lbl); ws.cell(row,6).value=v(key); ws.cell(row,7).value=f"=ROUND(F{row}*$F$3,0)"
    _c(ws,19,5,"V. 영업손익",bold=True,fill=TOTAL_FILL); ws.cell(19,6).value="=F11-F12"; ws.cell(19,7).value="=G11-G12"
    _c(ws,20,5,"VI. 영업외수익",bold=True,fill=TOTAL_FILL)
    ws.cell(20,6).value="=SUM(F21:F24)"; ws.cell(20,7).value="=SUM(G21:G24)"
    for row,key,lbl in [(21,"interest_income","1. 이자수익"),(22,"dividend_income","2. 배당금수익"),
        (23,"debt_forgiveness","3. 채무면제익"),(24,"other_non_op_income","4. 기타영업외수익")]:
        _c(ws,row,5,lbl); ws.cell(row,6).value=v(key); ws.cell(row,7).value=f"=ROUND(F{row}*$F$3,0)"
    _c(ws,25,5,"VII. 영업외비용",bold=True,fill=TOTAL_FILL)
    ws.cell(25,6).value="=SUM(F26:F27)"; ws.cell(25,7).value="=SUM(G26:G27)"
    for row,key,lbl in [(26,"interest_expense","1. 이자비용"),(27,"other_non_op_expense","2. 기타영업외비용")]:
        _c(ws,row,5,lbl); ws.cell(row,6).value=v(key); ws.cell(row,7).value=f"=ROUND(F{row}*$F$3,0)"
    _c(ws,28,5,"Ⅷ.법인세비용차감전손익",bold=True,fill=TOTAL_FILL)
    ws.cell(28,6).value="=F19+F20-F25"; ws.cell(28,7).value="=G19+G20-G25"
    _c(ws,29,5,"Ⅸ.법인세비용"); ws.cell(29,6).value=v("income_tax"); ws.cell(29,7).value="=ROUND(F29*$F$3,0)"
    _c(ws,30,5,"X. 당기순손익",bold=True,fill=TOTAL_FILL)
    ws.cell(30,6).value="=F28-F29"; ws.cell(30,7).value="=G28-G29"

    _c(ws,5,9,"I. 미처분이잉/결손금",bold=True,fill=TOTAL_FILL); ws.cell(5,10).value="=SUM(J6:J9)"
    _c(ws,6,9,"1. 전기이월미처분이잉/결손금"); ws.cell(6,10).value=(prior if prior else None)
    _c(ws,7,9,"2. 전기오류수정손익"); _c(ws,8,9,"3. 중간배당액")
    _c(ws,9,9,"4. 당기순손익"); ws.cell(9,10).value="=G30"
    _c(ws,10,9,"IV. 차기이월이잉(결손)",bold=True,fill=TOTAL_FILL); ws.cell(10,10).value="=J5"

    for col,w in [("B",33),("C",16),("D",16),("E",30),("F",16),("G",16),("H",3),("I",28),("J",16)]:
        ws.column_dimensions[col].width = w
    for row in range(3,34):
        for col in (3,6): ws.cell(row,col).number_format = "#,##0.00"
        for col in (4,7,10): ws.cell(row,col).number_format = "#,##0"


def convert(companies: list) -> bytes:
    import openpyxl as xl
    wb = xl.Workbook(); wb.remove(wb.active)

    for info in companies:
        bs_stream = info["bs_stream"]
        pl_stream = info.get("pl_stream")
        fmt       = info["fmt"]   # "excel" or "pdf"

        if fmt == "excel":
            bs, pl = parse_excel(bs_stream)
        else:
            raw = parse_pdf(bs_stream)
            bs, pl = _split_pdf_accounts(raw)

        mapping = build_mapping(bs, pl, info["company"], info["currency"])
        data    = apply_mapping(bs, pl, mapping)
        write_sheet(wb, info["company"], info["sheet_name"],
                    data, info["eoy_rate"], info["avg_rate"], info["prior_re"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
